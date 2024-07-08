import os
import time
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
import yfinance as yf
import json

API_KEY = '17b581337emsh2c49e9ba4d432a0p139b8ajsn3f1cd6b23669'


def read_tickers_from_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    tickers = []
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            tickers.append(cell.value)

    return tickers


file_path = 'company-reviews-test.xlsx'
tickers = read_tickers_from_excel(file_path)

companies = tickers


def fetch_adjusted_close_prices(symbol, start_date, end_date):
    try:
        ticker = yf.Ticker(symbol)
        data = ticker.history(interval='1mo', start=start_date, end=end_date)
        data = data['Close'].sort_index(ascending=True)
        return data
    except:
        return None

def fetch_daily_data(symbol, start_date, end_date):
    try:
        ticker = yf.Ticker(symbol)
        data = ticker.history(interval='1d', start=start_date, end=end_date)
        data = data[data.index.dayofweek < 5]  # Filter data for weekdays only
        data = data['Close'].sort_index(ascending=True)
        return data
    except:
        return None


def fetch_operating_income(symbol):
    url = "https://yh-finance.p.rapidapi.com/stock/v2/get-financials"
    querystring = {"symbol": symbol, "region": "US"}
    headers = {
        "X-RapidAPI-Key": API_KEY,
        "X-RapidAPI-Host": "yh-finance.p.rapidapi.com"
    }

    try:
        response = requests.request("GET", url, headers=headers, params=querystring)
        data = json.loads(response.text)
        income_statement_history = data['incomeStatementHistory']['incomeStatementHistory']

        operating_income = [
            income_statement_history[3]['operatingIncome']['raw'],
            income_statement_history[2]['operatingIncome']['raw'],
            income_statement_history[1]['operatingIncome']['raw']
        ]

        return operating_income
    except:
        return [None, None, None]


wb = Workbook()
ws = wb.active
ws.title = 'Adjusted Close Prices'

ws.cell(row=1, column=1, value='Symbol')

dates = pd.date_range(start='2018-12-31', end='2021-12-31', freq='MS')
col = 2
for i, date in enumerate(dates):
    ws.cell(row=1, column=col, value=f'Adj_Close({date.strftime("%Y-%m-%d")})')
    col += 1

for year in range(2019, 2022):
    ws.cell(row=1, column=col, value=f'Operating_{year}')
    col += 1

row = 2

for symbol in companies:
    print(f'Fetching data for {symbol}...')
    adjusted_close_prices = fetch_adjusted_close_prices(symbol, '2018-12-31', '2021-12-31')
    spring_daily_data = fetch_daily_data(symbol, '2020-03-02', '2020-03-20')
    winter_daily_data = fetch_daily_data(symbol, '2020-12-03', '2020-12-23')
    operating_income = fetch_operating_income(symbol)

    ws.cell(row=row, column=1, value=symbol)
    
    col = 2
    if adjusted_close_prices is not None and not adjusted_close_prices.empty:
        timezone = adjusted_close_prices.index.tz
        complete_date_range = pd.date_range(start='2018-12-31', end='2021-12-31', freq='MS').tz_localize(timezone)
        adjusted_close_prices = adjusted_close_prices.reindex(complete_date_range, method='ffill')
        
        for _, price in adjusted_close_prices.items():
            ws.cell(row=row, column=col, value=price)
            col += 1
    else:
        col += len(dates)

    if spring_daily_data is not None and not spring_daily_data.empty:
        timezone = spring_daily_data.index.tz
        complete_date_range = pd.date_range(start='2020-03-02', end='2020-03-20', freq='B').tz_localize(timezone)
        spring_daily_data = spring_daily_data.reindex(complete_date_range, method='ffill')
        
        for i, (_, price) in enumerate(spring_daily_data.items()):
            if row == 2:
                ws.cell(row=1, column=col, value=f'Spring_{i + 1}')
            ws.cell(row=row, column=col, value=price)
            col += 1
    else:
        col += len(pd.date_range(start='2020-03-02', end='2020-03-20', freq='B'))

    if winter_daily_data is not None and not winter_daily_data.empty:
        timezone = winter_daily_data.index.tz
        complete_date_range = pd.date_range(start='2020-12-03', end='2020-12-23', freq='B').tz_localize(timezone)
        winter_daily_data = winter_daily_data.reindex(complete_date_range, method='ffill')
        
        for i, (_, price) in enumerate(winter_daily_data.items()):
            if row == 2:
                ws.cell(row=1, column=col, value=f'Winter_{i + 1}')
            ws.cell(row=row, column=col, value=price)
            col += 1
    else:
        col += len(pd.date_range(start='2020-12-03', end='2020-12-23', freq='B'))

    for income in operating_income:
        ws.cell(row=row, column=col, value=income)
        col += 1

    row += 1




wb.save('AAAAAA.xlsx')
print('Saved data to adjusted_close_prices.xlsx')
