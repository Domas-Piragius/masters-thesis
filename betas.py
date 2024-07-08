import os
import time
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
import yfinance as yf
import json

API_KEY = '17b581337emsh2c49e9ba4d432a0p139b8ajsn3f1cd6b23669'


def read_tickers_from_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    tickers = []
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            tickers.append(cell.value)

    return tickers


file_path = 'company-reviews-test.xlsx'
tickers = read_tickers_from_excel(file_path)

companies = tickers


def fetch_daily_data(symbol, start_date, end_date):
    try:
        ticker = yf.Ticker(symbol)
        data = ticker.history(interval='1d', start=start_date, end=end_date)
        data = data[data.index.dayofweek < 5]  # Filter data for weekdays only
        data = data['Close'].sort_index(ascending=True)
        return data
    except:
        return None


wb = Workbook()
ws = wb.active
ws.title = 'Daily Stock Data'

ws.cell(row=1, column=1, value='Ticker')

date_range = pd.date_range(start='2019-06-01', end='2019-08-31', freq='B')
col = 2
for i, date in enumerate(date_range):
    ws.cell(row=1, column=col, value=f'Daily_{i + 1}')
    col += 1

row = 2

for symbol in companies:
    print(f'Fetching data for {symbol}...')
    daily_data = fetch_daily_data(symbol, '2019-06-01', '2019-08-31')

    ws.cell(row=row, column=1, value=symbol)

    col = 2
    if daily_data is not None and not daily_data.empty:
        timezone = daily_data.index.tz
        complete_date_range = pd.date_range(start='2019-06-01', end='2019-08-31', freq='B').tz_localize(timezone)
        daily_data = daily_data.reindex(complete_date_range, method='ffill')

        for _, price in daily_data.items():
            ws.cell(row=row, column=col, value=price)
            col += 1
    else:
        col += len(date_range)

    row += 1

# Add this function to your code to fetch market data
def fetch_market_data(start_date, end_date):
    try:
        market_ticker = '^GSPC'  # S&P 500
        ticker = yf.Ticker(market_ticker)
        data = ticker.history(interval='1d', start=start_date, end=end_date)
        data = data[data.index.dayofweek < 5]  # Filter data for weekdays only
        data = data['Close'].sort_index(ascending=True)
        return data
    except:
        return None

# Fetch market data before the loop
market_data = fetch_market_data('2019-06-01', '2019-08-31')

# Add market data to the output file
ws.cell(row=1, column=len(date_range) + 2, value='Market')
for i, (_, price) in enumerate(market_data.items()):
    ws.cell(row=i + 2, column=len(date_range) + 2, value=price)


# Save the output file with stock and market data
wb.save('daily_stock_and_market_data.xlsx')
print('Saved data to daily_stock_data.xlsx')
