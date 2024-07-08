import pandas as pd
import numpy as np
import openpyxl


def read_data_from_excel(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
    return df


sheet_name = 'Sheet1'
data = read_data_from_excel("daily_data.xlsx", sheet_name)

market_data = data.loc['SPY']
stock_data = data.drop('SPY')

stock_return = stock_data.pct_change(axis=1).dropna(axis=1, thresh=1)
market_returns = market_data.pct_change().dropna()

betas = []

for _, row in stock_return.iterrows():
    if row.isna().all():
        betas.append(None)
    else:
        cleaned_row = row.dropna()
        cleaned_market_returns = market_returns[cleaned_row.index]
        covariance_matrix = np.cov(cleaned_row, cleaned_market_returns, rowvar=False)
        beta = covariance_matrix[0, 1] / covariance_matrix[1, 1]
        betas.append(beta)

output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.cell(row=1, column=1, value="Ticker")
output_ws.cell(row=1, column=2, value="Beta")

for index, (ticker, beta) in enumerate(zip(stock_data.index, betas)):
    output_ws.cell(row=index + 2, column=1, value=ticker)
    output_ws.cell(row=index + 2, column=2, value=beta)

output_wb.save("betas.xlsx")
